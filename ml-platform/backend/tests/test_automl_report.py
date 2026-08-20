import shutil
import tempfile
import unittest
import uuid
import zipfile
from contextlib import contextmanager
from pathlib import Path
from types import SimpleNamespace

import joblib
import numpy as np
import pandas as pd
from pandas import read_csv
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.neighbors import KNeighborsClassifier

from app.services.automl_report import REPORT_FILES, REPORT_RESULT_COLUMNS, _importance, generate_automl_report


class FakeArtifactService:
    def __init__(self, dataset_path: Path, model_path: Path, output: Path):
        self.paths = {"dataset": dataset_path, "model": model_path}
        self.output = output
        self.created = {}

    def resolve(self, artifact_id, _project_id, expected_type=None):
        if str(artifact_id) not in self.paths and str(artifact_id) not in self.created:
            raise ValueError("missing")
        return SimpleNamespace(id=artifact_id, type=expected_type)

    @contextmanager
    def materialize(self, artifact_id, _project_id, expected_type=None):
        key = str(artifact_id)
        yield self.paths[key] if key in self.paths else self.created[key]

    def create_from_file(self, _project_id, source_path, name, _artifact_type, metadata=None):
        identifier = str(uuid.uuid4())
        destination = self.output / name
        shutil.copyfile(source_path, destination)
        self.created[identifier] = destination
        return SimpleNamespace(id=identifier)


class FakeDb:
    def __init__(self):
        self.commits = 0

    def commit(self):
        self.commits += 1


class AutoMLReportTests(unittest.TestCase):
    def test_importance_supports_coefficients_and_permutation_fallback(self):
        features = pd.DataFrame({"a": [-2, -1, 1, 2, 3, 4], "b": [0, 1, 0, 1, 0, 1]})
        target = pd.Series([0, 0, 1, 1, 1, 1])
        linear = LogisticRegression(random_state=42).fit(features, target)
        _, _, linear_weights = _importance(linear, features, target, "classification")
        self.assertAlmostEqual(float(linear_weights.sum()), 1.0)
        neighbor = KNeighborsClassifier(n_neighbors=1).fit(features, target)
        _, first_values, first_weights = _importance(neighbor, features, target, "classification")
        _, second_values, second_weights = _importance(neighbor, features, target, "classification")
        np.testing.assert_allclose(first_values, second_values)
        np.testing.assert_allclose(first_weights, second_weights)

    def test_generate_detailed_report_files_and_reuse_manifest(self):
        with tempfile.TemporaryDirectory() as temporary:
            tmp_path = Path(temporary)
            rng = np.random.default_rng(42)
            features = pd.DataFrame(rng.normal(size=(40, 3)), columns=["current", "pressure", "time"])
            target = np.where(features["current"] + features["pressure"] > 0, "defect", "ok")
            frame = features.assign(result=target)
            dataset_path = tmp_path / "dataset.csv"
            frame.to_csv(dataset_path, index=False)
            classes = sorted(set(target))
            encoded = np.array([classes.index(value) for value in target])
            model = RandomForestClassifier(n_estimators=20, random_state=42).fit(features, encoded)
            model_path = tmp_path / "model.joblib"
            joblib.dump({"model": model, "target_schema": {"task": "classification", "classes": classes}}, model_path)
            service = FakeArtifactService(dataset_path, model_path, tmp_path)
            job = SimpleNamespace(
                id=uuid.uuid4(), project_id=uuid.uuid4(), operator_id="automl", status="completed",
                dataset_artifact_id="dataset", model_artifact_id="model", name="report-job",
                params={"target_column": "result", "input_columns": list(features.columns), "task": "classification"},
                metrics={"best_model": {"name": "Random Forest"}, "all_results": [{"name": "Random Forest", "algorithm_id": "random_forest", "auc": 0.97, "f1": 0.94, "score": 0.95, "best_params": {"n_estimators": 20}, "training_time_seconds": 1.25, "status": "completed", "model_library_id": "library-1"}]},
                project=SimpleNamespace(name="project"), experiment=SimpleNamespace(name="experiment"),
            )
            db = FakeDb()

            manifest = generate_automl_report(db, job, service)
            with zipfile.ZipFile(tmp_path / "automl-report.zip") as archive:
                self.assertEqual(tuple(archive.namelist()), REPORT_FILES)
                archive.extract("automl_results.csv", tmp_path / "unzipped")
            report_results = read_csv(tmp_path / "unzipped" / "automl_results.csv")
            self.assertEqual(list(report_results.columns), REPORT_RESULT_COLUMNS)
            self.assertNotIn("score", report_results.columns)
            self.assertAlmostEqual(float(report_results.iloc[0]["Accuracy"]), 0.95)
            workbook = load_workbook(tmp_path / "AutoML全流程报告.xlsx")
            self.assertEqual(workbook.sheetnames, ["总览", "AutoML选型", "聚类画像", "特征重要性", "推理结果"])
            self.assertEqual(
                [cell.value for cell in workbook["AutoML选型"][1]],
                REPORT_RESULT_COLUMNS,
            )
            for sheet_name in ["总览", "聚类画像", "特征重要性"]:
                worksheet = workbook[sheet_name]
                self.assertEqual(len(worksheet._images), 1)
                self.assertGreater(worksheet._images[0].anchor._from.row + 1, worksheet.max_row)
            workbook.close()
            for filename in REPORT_FILES[2:]:
                self.assertTrue((tmp_path / filename).read_bytes().startswith(b"\x89PNG\r\n\x1a\n"))
            self.assertIn(manifest["preview"]["overview"]["best_k"], range(2, 9))
            self.assertEqual(len(manifest["preview"]["inference"]), len(frame))
            self.assertEqual(list(manifest["preview"]["inference"][0].keys())[:2], ["result", "predicted"])
            importance_values = [row["importance"] for row in manifest["preview"]["importance"]]
            self.assertEqual(importance_values, sorted(importance_values, reverse=True))
            created_count = len(service.created)
            self.assertEqual(generate_automl_report(db, job, service), manifest)
            self.assertEqual(len(service.created), created_count)
            generate_automl_report(db, job, service, regenerate=True)
            self.assertGreater(len(service.created), created_count)
