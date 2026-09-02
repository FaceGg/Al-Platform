import json
import unittest
import tempfile
import uuid
import warnings
from collections import Counter
from pathlib import Path
from unittest.mock import patch

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.metrics import f1_score, roc_auc_score
from sklearn.model_selection import StratifiedKFold
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base
from app.models.project import Project
from app.models.artifact import Artifact
from app.models.spot_weld_quality import (
    SpotWeldLabelSnapshot,
    SpotWeldQualityRuleSet,
    SpotWeldQualityRun,
    SpotWeldQualitySample,
)
from app.models.user import User
from app.services.artifact_service import ArtifactService
from app.services.automl_search import FamilySearchResult
from app.services.spot_weld_quality import (
    CandidateResult,
    apply_annotation_process_rules,
    assign_cluster_labels,
    _quality_estimator_metrics,
    apply_report_v1_rules,
    annotation_feature_frame,
    build_demo_report_frame,
    claim_quality_run,
    create_demo_quality_dataset,
    create_quality_run_record,
    execute_quality_run,
    read_report_dataset,
    recover_orphaned_local_quality_runs,
    run_automl,
    run_clustering,
    run_registered_model_annotation,
    run_snapshot_training,
    save_labeled_dataset,
    select_best_candidate,
    normalize_quality_search_config,
    normalize_annotation_process_rules,
    train_label_snapshot,
    update_quality_run_rules,
    validate_report_frame,
    warning_level,
)
from app.services.spot_weld_features import FEATURE_SCHEMA, QualityPipelineError, build_feature_frame
from app.storage.local import LocalStorage


class TestSpotWeldQualityService(unittest.TestCase):
    def test_annotation_feature_frame_materializes_legacy_report_model_features(self):
        frame = build_demo_report_frame(12)
        bundle = {"feature_schema_names": list(FEATURE_SCHEMA)}

        features = annotation_feature_frame(frame, bundle)

        self.assertEqual(list(features.columns), list(FEATURE_SCHEMA))
        self.assertEqual(features.shape, (12, len(FEATURE_SCHEMA)))
        self.assertTrue(np.isfinite(features.to_numpy(dtype=np.float64)).all())

    def test_automatic_rule_update_recalculates_labels_and_preserves_manual_labels(self):
        engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Session = sessionmaker(bind=engine)
        Base.metadata.create_all(engine)
        db = Session()
        try:
            owner = User(username=f"quality-rules-{uuid.uuid4().hex}", password_hash="hash")
            db.add(owner); db.flush()
            project = Project(name="Quality rules", owner_id=owner.id)
            db.add(project); db.flush()
            artifact = Artifact(
                project_id=project.id, name="rules.csv", type="dataset", storage_path="rules.csv",
                format="csv", file_size=1, metadata_={"row_count": 2},
            )
            db.add(artifact); db.flush()
            run = SpotWeldQualityRun(
                project_id=project.id, dataset_artifact_id=artifact.id, created_by_id=owner.id,
                status="completed", input_fingerprint={"label_mode": "automatic", "rule_config": {}},
                statistics={"annotation_progress": {"annotated_count": 2, "total_count": 2, "percent": 100.0}},
            )
            db.add(run); db.flush()
            samples = [
                SpotWeldQualitySample(
                    run_id=run.id, source_row_index=0, display_id="W-0001",
                    feature_values={"wld_spatter_strength": 3, "spotdiameter": 5, "energy_dev": 0,
                                    "current_max_diff": 1, "power_std": 1},
                    automatic_label="strong_splatter", current_label="spot_too_small",
                    rule_hits=[{"code": "strong_splatter"}], cluster_id=0,
                ),
                SpotWeldQualitySample(
                    run_id=run.id, source_row_index=1, display_id="W-0002",
                    feature_values={"wld_spatter_strength": 1, "spotdiameter": 5, "energy_dev": 0,
                                    "current_max_diff": 2, "power_std": 2},
                    automatic_label="normal", current_label=None, rule_hits=[], cluster_id=0,
                ),
            ]
            db.add_all(samples); db.commit()
            progress = []

            updated = update_quality_run_rules(
                db,
                run,
                {"strong_splatter_min": 4, "weak_splatter_value": 2},
                batch_size=1,
                progress_callback=lambda value: progress.append(value["annotated_count"]),
            )

            db.refresh(samples[0]); db.refresh(samples[1])
            self.assertEqual(updated.status, "completed")
            self.assertEqual(samples[0].automatic_label, "normal")
            self.assertEqual(samples[0].current_label, "spot_too_small")
            self.assertEqual(progress, [0, 1, 2])
            self.assertEqual(updated.input_fingerprint["rule_config"]["strong_splatter_min"], 4)
        finally:
            db.close(); engine.dispose()

    def test_manual_rule_update_saves_configuration_without_changing_labels(self):
        engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Session = sessionmaker(bind=engine)
        Base.metadata.create_all(engine)
        db = Session()
        try:
            owner = User(username=f"quality-manual-rules-{uuid.uuid4().hex}", password_hash="hash")
            db.add(owner); db.flush()
            project = Project(name="Manual rules", owner_id=owner.id)
            db.add(project); db.flush()
            artifact = Artifact(project_id=project.id, name="rules.csv", type="dataset", storage_path="rules.csv", format="csv", file_size=1)
            db.add(artifact); db.flush()
            run = SpotWeldQualityRun(
                project_id=project.id, dataset_artifact_id=artifact.id, created_by_id=owner.id,
                status="completed", input_fingerprint={"label_mode": "manual", "rule_config": {}}, statistics={},
            )
            db.add(run); db.flush()
            sample = SpotWeldQualitySample(
                run_id=run.id, source_row_index=0, display_id="W-0001",
                feature_values={"wld_spatter_strength": 3}, automatic_label=None,
                current_label="weak_splatter", rule_hits=[],
            )
            db.add(sample); db.commit()

            updated = update_quality_run_rules(db, run, {"strong_splatter_min": 4})

            db.refresh(sample)
            self.assertEqual(updated.status, "completed")
            self.assertEqual(updated.input_fingerprint["rule_config"]["strong_splatter_min"], 4)
            self.assertIsNone(sample.automatic_label)
            self.assertEqual(sample.current_label, "weak_splatter")
        finally:
            db.close(); engine.dispose()

    def test_snapshot_candidate_metrics_use_macro_multiclass_auc_and_f1(self):
        class BiasedClassifier:
            feature_importances_ = np.ones(3, dtype=float)

            def fit(self, _features, _target):
                return self

            @staticmethod
            def predict_proba(features):
                encoded = np.asarray(features, dtype=float).argmax(axis=1)
                probabilities = np.full((len(encoded), 3), 0.05, dtype=float)
                probabilities[encoded != 2, 0] = 0.9
                probabilities[encoded == 2, 2] = 0.9
                return probabilities

            def predict(self, features):
                return self.predict_proba(features).argmax(axis=1)

        raw_labels = np.asarray([0] * 10 + [1] * 5 + [2] * 5)
        features = np.eye(3, dtype=float)[raw_labels]
        labels = np.asarray(["normal", "weak_splatter", "strong_splatter"])[raw_labels]
        _, encoded_labels = np.unique(labels, return_inverse=True)
        macro_auc_values = []
        weighted_auc_values = []
        macro_f1_values = []
        weighted_f1_values = []
        for _, test_index in StratifiedKFold(n_splits=5, shuffle=True, random_state=42).split(features, encoded_labels):
            probabilities = BiasedClassifier.predict_proba(features[test_index])
            predictions = probabilities.argmax(axis=1)
            macro_auc_values.append(roc_auc_score(
                encoded_labels[test_index], probabilities, labels=np.arange(3), multi_class="ovr", average="macro",
            ))
            weighted_auc_values.append(roc_auc_score(
                encoded_labels[test_index], probabilities, labels=np.arange(3), multi_class="ovr", average="weighted",
            ))
            macro_f1_values.append(f1_score(encoded_labels[test_index], predictions, average="macro", zero_division=0))
            weighted_f1_values.append(f1_score(encoded_labels[test_index], predictions, average="weighted", zero_division=0))

        metrics = _quality_estimator_metrics(
            BiasedClassifier,
            features=features,
            target=encoded_labels,
            evaluation={"cross_validation_enabled": True, "cross_validation_folds": 5},
        )

        self.assertAlmostEqual(metrics["auc"], float(np.mean(macro_auc_values)))
        self.assertAlmostEqual(metrics["f1"], float(np.mean(macro_f1_values)))
        self.assertNotAlmostEqual(metrics["auc"], float(np.mean(weighted_auc_values)))
        self.assertNotAlmostEqual(metrics["f1"], float(np.mean(weighted_f1_values)))

    def test_save_labeled_dataset_appends_label_column_and_preserves_source_rows(self):
        with tempfile.TemporaryDirectory(prefix="quality-save-labels-") as directory:
            engine = create_engine(
                "sqlite://",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            Session = sessionmaker(bind=engine)
            Base.metadata.create_all(engine)
            db = Session()
            try:
                owner = User(username=f"quality-save-{uuid.uuid4().hex}", password_hash="hash")
                db.add(owner)
                db.flush()
                project = Project(name="Quality save", owner_id=owner.id)
                db.add(project)
                db.commit()
                artifacts = ArtifactService(db, LocalStorage(Path(directory) / "artifacts"))
                dataset = create_demo_quality_dataset(
                    db,
                    project_id=project.id,
                    row_count=24,
                    artifact_service=artifacts,
                )
                db.commit()
                run = create_quality_run_record(
                    db,
                    project_id=project.id,
                    user_id=owner.id,
                    dataset_artifact_id=dataset.id,
                    algorithm_ids=["gbdt"],
                    search_method="random",
                    max_trials=5,
                    time_budget=60,
                    artifact_service=artifacts,
                )
                db.commit()
                self.assertEqual(execute_quality_run(db, run.id, artifact_service=artifacts).status, "completed")
                samples = db.query(SpotWeldQualitySample).filter(
                    SpotWeldQualitySample.run_id == run.id,
                ).order_by(SpotWeldQualitySample.source_row_index).all()
                samples[0].current_label = "normal"
                samples[1].current_label = "spot_too_small"
                db.commit()

                labeled = save_labeled_dataset(db, run, artifact_service=artifacts)
                db.commit()

                self.assertNotEqual(labeled.id, dataset.id)
                self.assertEqual(labeled.type, "dataset")
                self.assertEqual(labeled.metadata_["source_dataset_artifact_id"], str(dataset.id))
                with artifacts.materialize(labeled.id, project.id, expected_type="dataset") as path:
                    frame = read_report_dataset(path)
                self.assertEqual(len(frame), 24)
                self.assertEqual(list(frame.columns)[-1], "label")
                self.assertEqual(frame.loc[0, "label"], "normal")
                self.assertEqual(frame.loc[1, "label"], "spot_too_small")
                self.assertTrue(frame["label"].notna().all())
            finally:
                db.close()
                engine.dispose()

    def test_save_labeled_dataset_uses_selected_target_name_and_dtype(self):
        with tempfile.TemporaryDirectory(prefix="quality-save-typed-labels-") as directory:
            engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
            Session = sessionmaker(bind=engine)
            Base.metadata.create_all(engine)
            db = Session()
            try:
                owner = User(username=f"typed-labels-{uuid.uuid4().hex}", password_hash="hash")
                db.add(owner); db.flush()
                project = Project(name="Typed labels", owner_id=owner.id)
                db.add(project); db.flush()
                artifacts = ArtifactService(db, LocalStorage(Path(directory) / "artifacts"))
                frame = build_demo_report_frame(12)
                frame["Fault"] = np.tile([0, 1], 6)
                source_path = Path(directory) / "fault.csv"
                frame.to_csv(source_path, index=False)
                dataset = artifacts.create_from_file(project.id, source_path, "fault.csv", "dataset")
                db.flush()
                run = SpotWeldQualityRun(
                    project_id=project.id,
                    dataset_artifact_id=dataset.id,
                    created_by_id=owner.id,
                    status="completed",
                    input_fingerprint={
                        "label_mode": "manual",
                        "target_column": "Fault",
                        "target_column_created": False,
                        "target_schema": {"name": "Fault", "dtype": "int64", "classes": ["0", "1"]},
                    },
                    statistics={},
                )
                db.add(run); db.flush()
                db.add_all([
                    SpotWeldQualitySample(run_id=run.id, source_row_index=index, display_id=f"W-{index + 1:04d}", current_label=str(index % 2))
                    for index in range(len(frame))
                ])
                db.commit()
                labeled = save_labeled_dataset(db, run, artifact_service=artifacts)
                db.commit()
                with artifacts.materialize(labeled.id, project.id, expected_type="dataset") as path:
                    saved = read_report_dataset(path)
                self.assertEqual(saved.columns[-1], "Fault")
                self.assertEqual(str(saved["Fault"].dtype), "int64")
                self.assertEqual(saved["Fault"].tolist()[:4], [0, 1, 0, 1])
            finally:
                db.close(); engine.dispose()

    def test_quality_search_config_preserves_order_and_rejects_invalid_sets(self):
        selected = normalize_quality_search_config(
            ["random_forest", "gbdt"], "random", 5, 60,
        )
        self.assertEqual(selected["algorithm_ids"], ["random_forest", "gbdt"])
        self.assertEqual(len(normalize_quality_search_config([], "bayesian", 20, 600)["algorithm_ids"]), 7)
        for algorithm_ids in (["random_forest", "random_forest"], ["does-not-exist"]):
            with self.subTest(algorithm_ids=algorithm_ids):
                with self.assertRaises(QualityPipelineError) as raised:
                    normalize_quality_search_config(algorithm_ids, "random", 5, 60)
                self.assertEqual(raised.exception.code, "QUALITY_AUTOML_SEARCH_CONFIG_INVALID")

    def test_candidate_selection_orders_auc_then_f1_then_index(self):
        results = [
            CandidateResult("lightgbm", "LightGBM", "completed", 0, auc=0.91, f1=0.80),
            CandidateResult("catboost", "CatBoost", "completed", 1, auc=0.91, f1=0.82),
            CandidateResult("random_forest", "Random Forest", "completed", 2, auc=0.91, f1=0.82),
        ]
        self.assertEqual(select_best_candidate(results).algorithm_id, "catboost")

    def test_automl_searches_selected_algorithm_families(self):
        features = np.asarray([
            [index % 11, (index * 3) % 7, index % 5]
            for index in range(60)
        ], dtype=float)
        labels = np.asarray([
            int((index % 11) + ((index * 3) % 7) > 8)
            for index in range(60)
        ])

        results, winner = run_automl(
            features,
            labels,
            algorithm_ids=["gbdt", "random_forest"],
            search_method="bayesian",
            max_trials=5,
            time_budget=60,
            evaluation={"cross_validation_enabled": True, "cross_validation_folds": 3},
        )

        self.assertEqual([item.algorithm_id for item in results], ["gbdt", "random_forest"])
        self.assertIn(winner.algorithm_id, {"gbdt", "random_forest"})
        self.assertTrue(winner.best_params)
        self.assertIsNotNone(winner.auc)
        self.assertIsNotNone(winner.f1)

    def test_automl_continues_after_an_unavailable_family(self):
        features = np.asarray([
            [index % 11, (index * 3) % 7, index % 5]
            for index in range(60)
        ], dtype=float)
        labels = np.asarray([
            int((index % 11) + ((index * 3) % 7) > 8)
            for index in range(60)
        ])

        def family_search(**kwargs):
            family = kwargs["family"]
            if family.id == "gbdt":
                return FamilySearchResult(
                    algorithm_id=family.id,
                    display_name=family.display_name,
                    catalog_index=kwargs["catalog_index"],
                    status="unavailable",
                    error_code="AUTOML_ALGORITHM_UNAVAILABLE",
                )
            return FamilySearchResult(
                algorithm_id=family.id,
                display_name=family.display_name,
                catalog_index=kwargs["catalog_index"],
                status="completed",
                best_score=0.8,
                best_params=dict(family.default_params),
                completed_trials=1,
            )

        results, winner = run_automl(
            features,
            labels,
            algorithm_ids=["gbdt", "random_forest"],
            search_method="random",
            max_trials=5,
            time_budget=60,
            evaluation={"cross_validation_enabled": False, "cross_validation_folds": None},
            family_search=family_search,
        )

        self.assertEqual(results[0].status, "unavailable")
        self.assertEqual(results[0].error_code, "AUTOML_ALGORITHM_UNAVAILABLE")
        self.assertEqual(winner.algorithm_id, "random_forest")

    def test_report_rules_keep_all_hits_in_table_order(self):
        result = apply_report_v1_rules(
            {"wld_spatter_strength": 3, "power_std": 99, "spotdiameter": 5},
            thresholds={"power_std_p95": 1},
        )
        self.assertEqual(result.primary_label, "strong_splatter")
        self.assertEqual([hit.code for hit in result.hits], ["strong_splatter", "power_fluctuation"])

    def test_report_rules_follow_configured_thresholds_and_fixed_cluster_one(self):
        thresholds = {
            "energy_dev_sigma": 2.5,
            "current_max_diff_p95": 100.0,
            "power_std_p95": 20.0,
        }

        cases = (
            ({"wld_spatter_strength": 3}, None, None, ["strong_splatter"]),
            ({"wld_spatter_strength": 2}, None, None, ["weak_splatter"]),
            ({"spotdiameter": 1.99}, None, None, ["spot_too_small"]),
            ({"spotdiameter": 80.01}, None, None, ["spot_too_large"]),
            ({"energy_dev": -2.51}, None, None, ["energy_anomaly"]),
            ({"current_max_diff": 100.01}, None, None, ["current_jump"]),
            ({"power_std": 20.01}, None, None, ["power_fluctuation"]),
            ({"wld_spatter_strength": 2}, 1, 0, ["weak_splatter", "anomaly_cluster"]),
            ({}, None, None, ["normal"]),
        )

        for values, cluster_id, anomaly_cluster, expected_codes in cases:
            with self.subTest(values=values, cluster_id=cluster_id):
                result = apply_report_v1_rules(
                    values,
                    thresholds=thresholds,
                    cluster_id=cluster_id,
                    anomaly_cluster=anomaly_cluster,
                )
                self.assertEqual(result.hit_codes, expected_codes)

        self.assertNotIn(
            "energy_anomaly",
            apply_report_v1_rules({"energy_dev": 2.5}, thresholds=thresholds).hit_codes,
        )
        self.assertNotIn(
            "current_jump",
            apply_report_v1_rules({"current_max_diff": 100.0}, thresholds=thresholds).hit_codes,
        )
        self.assertNotIn(
            "power_fluctuation",
            apply_report_v1_rules({"power_std": 20.0}, thresholds=thresholds).hit_codes,
        )
        self.assertNotIn(
            "spot_too_small",
            apply_report_v1_rules({"spotdiameter": 0.0}, thresholds=thresholds).hit_codes,
        )
        self.assertNotIn(
            "spot_too_small",
            apply_report_v1_rules({"spotdiameter": 2.0}, thresholds=thresholds).hit_codes,
        )
        self.assertNotIn(
            "spot_too_large",
            apply_report_v1_rules({"spotdiameter": 80.0}, thresholds=thresholds).hit_codes,
        )
        self.assertNotIn(
            "anomaly_cluster",
            apply_report_v1_rules(
                {"wld_spatter_strength": 2},
                thresholds=thresholds,
                cluster_id=0,
                anomaly_cluster=1,
            ).hit_codes,
        )

    def test_report_rules_allow_the_annotation_editor_to_override_all_rule_values(self):
        thresholds = {
            "strong_splatter_min": 4,
            "weak_splatter_value": 1,
            "spotdiameter_small_min": 0.5,
            "spotdiameter_small_max": 1.5,
            "spotdiameter_large_min": 90,
            "energy_dev_sigma": 3.0,
            "current_max_diff_p95": 120.0,
            "power_std_p95": 30.0,
            "spatter_cluster_id": 2,
            "spatter_cluster_min_strength": 3,
        }

        self.assertEqual(
            apply_report_v1_rules({"wld_spatter_strength": 3}, thresholds=thresholds).primary_label,
            "normal",
        )
        self.assertEqual(
            apply_report_v1_rules({"wld_spatter_strength": 1}, thresholds=thresholds).primary_label,
            "weak_splatter",
        )
        self.assertEqual(
            apply_report_v1_rules({"spotdiameter": 1.0}, thresholds=thresholds).primary_label,
            "spot_too_small",
        )
        self.assertEqual(
            apply_report_v1_rules({"spotdiameter": 80.1}, thresholds=thresholds).primary_label,
            "normal",
        )
        self.assertEqual(
            apply_report_v1_rules({"wld_spatter_strength": 3}, thresholds=thresholds, cluster_id=2).hit_codes,
            ["anomaly_cluster"],
        )

    def test_zero_spot_diameter_is_not_virtual_weld(self):
        result = apply_report_v1_rules({"spotdiameter": 0}, thresholds={})
        self.assertNotIn("spot_too_small", result.hit_codes)

    def test_warning_probability_maps_to_four_levels(self):
        self.assertEqual(warning_level(0.8), "critical")
        self.assertEqual(warning_level(0.6), "warning")
        self.assertEqual(warning_level(0.3), "notice")
        self.assertEqual(warning_level(0.299), "none")

    def test_clustering_searches_k_and_returns_pca_coordinates(self):
        features = np.vstack([
            np.random.default_rng(42).normal(0, 0.2, (8, 3)),
            np.random.default_rng(7).normal(4, 0.2, (8, 3)),
        ])
        result = run_clustering(
            features,
            feature_names=["power_std", "energy_dev", "spatter_total"],
            feature_importance=np.array([3.0, 1.0, 2.0]),
        )
        self.assertIn(result.best_k, range(2, 9))
        self.assertEqual(len(result.cluster_ids), len(features))
        self.assertEqual(result.pca_coordinates.shape, (16, 2))

    def test_clustering_ignores_tiny_negative_importance_residuals(self):
        features = np.vstack([
            np.random.default_rng(42).normal(0, 0.2, (8, 3)),
            np.random.default_rng(7).normal(4, 0.2, (8, 3)),
        ])

        with warnings.catch_warnings(record=True) as caught:
            warnings.simplefilter("always")
            result = run_clustering(
                features,
                feature_names=["power_std", "energy_dev", "spatter_total"],
                feature_importance=np.array([1.0, -1e-12, 2.0]),
            )

        self.assertTrue(np.isfinite(result.pca_coordinates).all())
        self.assertTrue(np.isfinite(result.weights).all())
        self.assertTrue(all(weight >= 0 for weight in result.weights))
        self.assertFalse(any("invalid value encountered in sqrt" in str(item.message) for item in caught))

    def test_annotation_rules_consume_short_circuited_boolean_operands(self):
        rules = normalize_annotation_process_rules(
            [{
                "id": "rule-1",
                "label": "1",
                "tokens": [
                    {"kind": "data", "value": "temperature"},
                    {"kind": "logical_operator", "value": ">"},
                    {"kind": "number", "value": 10},
                    {"kind": "logical_operator", "value": "or"},
                    {"kind": "data", "value": "pressure"},
                    {"kind": "logical_operator", "value": ">"},
                    {"kind": "number", "value": 5},
                ],
            }],
            columns=["temperature", "pressure"],
            label_dtype="int",
        )

        label, hits = apply_annotation_process_rules({"temperature": 12, "pressure": 0}, rules)

        self.assertEqual(label, "1")
        self.assertEqual(hits[0]["code"], "rule-1")

    def test_annotation_rules_validate_columns_and_shared_label_dtype(self):
        base_rule = {
            "id": "rule-1",
            "tokens": [
                {"kind": "data", "value": "temperature"},
                {"kind": "logical_operator", "value": ">"},
                {"kind": "number", "value": 10},
            ],
        }
        with self.assertRaisesRegex(QualityPipelineError, "QUALITY_LABEL_TYPE_INVALID"):
            normalize_annotation_process_rules(
                [{**base_rule, "label": "not-an-int"}],
                columns=["temperature"],
                label_dtype="int",
            )
        with self.assertRaisesRegex(QualityPipelineError, "QUALITY_ANNOTATION_RULE_COLUMN_INVALID"):
            normalize_annotation_process_rules(
                [{**base_rule, "label": "1"}],
                columns=["pressure"],
                label_dtype="int",
            )

    def test_annotation_rules_use_fallback_only_after_condition_rules_miss(self):
        rules = normalize_annotation_process_rules(
            [
                {
                    "id": "hot",
                    "label": "hot",
                    "tokens": [
                        {"kind": "data", "value": "temperature"},
                        {"kind": "logical_operator", "value": ">"},
                        {"kind": "number", "value": 10},
                    ],
                },
                {"id": "other", "kind": "fallback", "label": "other", "tokens": []},
            ],
            columns=["temperature"],
            label_dtype="string",
        )

        self.assertEqual(rules[0]["kind"], "condition")
        self.assertEqual(rules[1]["kind"], "fallback")
        self.assertEqual(apply_annotation_process_rules({"temperature": 12}, rules)[0], "hot")
        fallback_label, fallback_hits = apply_annotation_process_rules({"temperature": 2}, rules)
        self.assertEqual(fallback_label, "other")
        self.assertEqual(fallback_hits[0]["code"], "other")

    def test_annotation_rules_reject_invalid_fallback_contracts(self):
        fallback = {"id": "other", "kind": "fallback", "label": "other", "tokens": []}
        invalid_rules = [
            [fallback, {**fallback, "id": "other-2"}],
            [{**fallback, "kind": "unknown"}],
            [{**fallback, "tokens": [{"kind": "number", "value": 1}]}],
        ]
        for rules in invalid_rules:
            with self.subTest(rules=rules), self.assertRaisesRegex(
                QualityPipelineError, "QUALITY_ANNOTATION_RULE_INVALID"
            ):
                normalize_annotation_process_rules(rules, columns=["temperature"], label_dtype="string")

        self.assertEqual(apply_annotation_process_rules({"temperature": 2}, []), (None, []))

    def test_registered_model_annotation_requires_valid_feature_importance(self):
        class ModelWithoutImportance:
            classes_ = np.array(["normal", "strong_splatter"])

            def predict(self, values):
                return np.zeros(len(values), dtype=int)

        features = pd.DataFrame(
            np.ones((3, 73), dtype=float),
            columns=list(FEATURE_SCHEMA),
        )

        with self.assertRaises(QualityPipelineError) as caught:
            run_registered_model_annotation(features, {
                "model": ModelWithoutImportance(),
                "feature_schema": list(FEATURE_SCHEMA),
                "classes": ["normal", "strong_splatter"],
            })

        self.assertEqual(caught.exception.code, "QUALITY_MODEL_FEATURE_IMPORTANCE_INVALID")

    def test_cluster_label_assignment_requires_complete_unique_single_labels(self):
        with self.assertRaises(QualityPipelineError) as caught:
            assign_cluster_labels([0, 1, 0], {"0": "normal"})
        self.assertEqual(caught.exception.code, "QUALITY_CLUSTER_LABELS_REQUIRED")

        with self.assertRaises(QualityPipelineError) as caught:
            assign_cluster_labels([0, 1, 0], {"0": "normal", "1": "normal"})
        self.assertEqual(caught.exception.code, "QUALITY_CLUSTER_LABELS_REQUIRED")

        self.assertEqual(
            assign_cluster_labels([0, 1, 0], {"0": "normal", "1": "strong_splatter"}),
            ["normal", "strong_splatter", "normal"],
        )

    def test_demo_dataset_is_report_compatible_and_has_multiple_label_groups(self):
        frame = build_demo_report_frame(24)
        features, schema, _ = build_feature_frame(frame)
        labels = [
            apply_report_v1_rules(record, thresholds={}).primary_label
            for record in features.to_dict(orient="records")
        ]

        self.assertEqual(len(frame), 24)
        self.assertEqual(len(schema), 73)
        self.assertGreaterEqual(sum(label == "normal" for label in labels), 3)
        self.assertGreaterEqual(sum(label != "normal" for label in labels), 3)

    def test_report_reproduction_demo_has_enough_automatic_labels_for_five_fold_training(self):
        frame = build_demo_report_frame(1875)
        features, _, _ = build_feature_frame(frame)
        thresholds = {
            "energy_dev_sigma": 2.5,
            "current_max_diff_p95": float(np.percentile(features["current_max_diff"], 95)),
            "power_std_p95": float(np.percentile(features["power_std"], 95)),
        }
        labels = [
            apply_report_v1_rules(record, thresholds=thresholds).primary_label
            for record in features.to_dict(orient="records")
        ]

        self.assertGreaterEqual(min(Counter(labels).values()), 5)

    def test_validation_counts_only_rows_that_fail_feature_extraction(self):
        frame = build_demo_report_frame(12)
        frame.loc[1, "cvei"] = "not-a-waveform"

        validation = validate_report_frame(frame)

        self.assertEqual(validation["row_count"], 12)
        self.assertEqual(validation["valid_rows"], 11)
        self.assertEqual(validation["invalid_rows"], 1)
        self.assertEqual(validation["errors"][0]["code"], "QUALITY_WAVEFORM_INVALID_BASE64")
        self.assertEqual(validation["errors"][0]["row_index"], 1)

    def test_automl_uses_matching_feature_names_for_lightgbm_prediction(self):
        features, _, _ = build_feature_frame(build_demo_report_frame(12))
        labels = np.asarray([
            apply_report_v1_rules(record, thresholds={}).primary_label != "normal"
            for record in features.to_dict(orient="records")
        ], dtype=int)

        with warnings.catch_warnings(record=True) as captured:
            warnings.simplefilter("always")
            run_automl(
                features.to_numpy(dtype=np.float64),
                labels,
                algorithm_ids=["lightgbm"],
                search_method="random",
                max_trials=5,
                time_budget=60,
            )

        self.assertFalse(any("does not have valid feature names" in str(item.message) for item in captured))

    def test_legacy_xls_parser_errors_are_stable_quality_errors(self):
        with tempfile.TemporaryDirectory(prefix="quality-xls-test-") as directory:
            path = Path(directory) / "report.xls"
            path.write_bytes(b"not-a-spreadsheet")
            with patch("app.services.spot_weld_quality.pd.read_excel", side_effect=ImportError("xlrd missing")):
                with self.assertRaises(QualityPipelineError) as raised:
                    read_report_dataset(path)

        self.assertEqual(raised.exception.code, "QUALITY_DATASET_INVALID")

    def test_atomic_claim_allows_only_one_worker_to_start_a_quality_run(self):
        engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Session = sessionmaker(bind=engine)
        Base.metadata.create_all(engine)
        db = Session()
        try:
            run = SpotWeldQualityRun(
                project_id=uuid.uuid4(),
                dataset_artifact_id=uuid.uuid4(),
                created_by_id=uuid.uuid4(),
                status="queued",
            )
            db.add(run)
            db.commit()

            first = claim_quality_run(db, run.id, worker_id="worker-a", task_id="task-a")
            second = claim_quality_run(db, run.id, worker_id="worker-b", task_id="task-b")

            self.assertIsNotNone(first)
            self.assertIsNone(second)
            db.refresh(run)
            self.assertEqual(run.status, "running")
            self.assertEqual(run.worker_id, "worker-a")
            self.assertEqual(run.task_id, "task-a")
        finally:
            db.close()
            engine.dispose()

    def test_local_run_recovery_marks_pre_restart_non_terminal_runs_failed(self):
        engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
        Session = sessionmaker(bind=engine)
        Base.metadata.create_all(engine)
        db = Session()
        try:
            local_runs = [
                SpotWeldQualityRun(
                    project_id=uuid.uuid4(),
                    dataset_artifact_id=uuid.uuid4(),
                    created_by_id=uuid.uuid4(),
                    status=status,
                    task_id=f"local:quality-{status}",
                    worker_id="local",
                )
                for status in ("queued", "validating", "running")
            ]
            external_run = SpotWeldQualityRun(
                project_id=uuid.uuid4(),
                dataset_artifact_id=uuid.uuid4(),
                created_by_id=uuid.uuid4(),
                status="running",
                task_id="celery:quality-running",
                worker_id="worker-a",
            )
            completed_local_run = SpotWeldQualityRun(
                project_id=uuid.uuid4(),
                dataset_artifact_id=uuid.uuid4(),
                created_by_id=uuid.uuid4(),
                status="completed",
                task_id="local:quality-completed",
                worker_id="local",
            )
            db.add_all([*local_runs, external_run, completed_local_run])
            db.commit()

            recovered = recover_orphaned_local_quality_runs(db)

            self.assertEqual(recovered, 3)
            db.expire_all()
            for run in local_runs:
                stored = db.get(SpotWeldQualityRun, run.id)
                self.assertEqual(stored.status, "failed")
                self.assertEqual(stored.error_code, "QUALITY_RUN_LOCAL_WORKER_RESTARTED")
                self.assertEqual(stored.error_details, {
                    "code": "QUALITY_RUN_LOCAL_WORKER_RESTARTED",
                    "message": "Local quality worker stopped during service restart; rerun this task.",
                })
            self.assertEqual(db.get(SpotWeldQualityRun, external_run.id).status, "running")
            self.assertEqual(db.get(SpotWeldQualityRun, completed_local_run.id).status, "completed")
        finally:
            db.close()
            engine.dispose()

    def test_demo_quality_run_persists_samples_and_generated_artifacts(self):
        with tempfile.TemporaryDirectory(prefix="quality-service-test-") as directory:
            engine = create_engine(
                "sqlite://",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            Session = sessionmaker(bind=engine)
            Base.metadata.create_all(engine)
            db = Session()
            try:
                owner = User(username=f"quality-service-{uuid.uuid4().hex}", password_hash="hash")
                db.add(owner)
                db.flush()
                project = Project(name="Quality service", owner_id=owner.id)
                db.add(project)
                db.commit()
                artifacts = ArtifactService(db, LocalStorage(Path(directory) / "artifacts"))
                dataset = create_demo_quality_dataset(
                    db,
                    project_id=project.id,
                    row_count=24,
                    artifact_service=artifacts,
                )
                db.commit()
                run = create_quality_run_record(
                    db,
                    project_id=project.id,
                    user_id=owner.id,
                    dataset_artifact_id=dataset.id,
                    algorithm_ids=["gbdt"],
                    search_method="random",
                    max_trials=5,
                    time_budget=60,
                    artifact_service=artifacts,
                )
                # Queued runs created before the report_v2 rules must record the
                # rules actually applied when execution begins.
                run.rule_set_version = "report_v1"
                db.commit()

                committed_annotation_counts = []

                def capture_annotation_progress(session):
                    progress = (run.statistics or {}).get("annotation_progress") or {}
                    if "annotated_count" in progress:
                        committed_annotation_counts.append(progress["annotated_count"])

                event.listen(db, "before_commit", capture_annotation_progress)

                with warnings.catch_warnings(record=True) as caught:
                    warnings.simplefilter("always", DeprecationWarning)
                    try:
                        outcome = execute_quality_run(db, run.id, artifact_service=artifacts)
                    finally:
                        event.remove(db, "before_commit", capture_annotation_progress)
                db.refresh(run)

                self.assertEqual(outcome.status, "completed")
                self.assertEqual(run.status, "completed")
                self.assertFalse(any(
                    "Conversion of an array with ndim" in str(item.message)
                    for item in caught
                ))
                self.assertEqual(run.rule_set_version, "report_v2")
                self.assertTrue(any(0 < count < 24 for count in committed_annotation_counts))
                self.assertEqual(
                    db.query(SpotWeldQualitySample).filter(SpotWeldQualitySample.run_id == run.id).count(),
                    24,
                )
                self.assertEqual(len(run.automl_results), 1)
                self.assertEqual(run.automl_results[0]["algorithm_id"], "gbdt")
                self.assertFalse(any(result["error_code"] for result in run.automl_results))
                self.assertEqual(
                    set(run.output_artifacts),
                    {
                        "features", "results", "report",
                        "model_comparison_chart", "cluster_pca_chart", "feature_importance_chart",
                        "warning_distribution_chart", "waveform_comparison_chart",
                    },
                )
                self.assertTrue(all(
                    float(result["training_time_seconds"]) > 0
                    for result in run.automl_results
                ))
                report_artifact = db.query(Artifact).filter(
                    Artifact.id == uuid.UUID(run.output_artifacts["report"]),
                ).one()
                self.assertEqual(report_artifact.type, "quality_report")
                self.assertEqual(report_artifact.metadata_["rule_set_version"], "report_v2")
                candidate_results = report_artifact.metadata_["candidate_results"]
                self.assertEqual(len(candidate_results), 1)
                self.assertEqual(
                    [item["name"] for item in candidate_results],
                    [item["name"] for item in run.automl_results],
                )
                with artifacts.materialize(
                    run.output_artifacts["results"],
                    project.id,
                    expected_type="quality_results",
                ) as results_path:
                    serialized_results = json.loads(results_path.read_text(encoding="utf-8"))["candidate_results"]
                self.assertEqual(len(serialized_results), 1)
                self.assertEqual(
                    [item["name"] for item in serialized_results],
                    [item["name"] for item in run.automl_results],
                )
                with artifacts.materialize(
                    run.output_artifacts["report"],
                    project.id,
                    expected_type="quality_report",
                ) as report_path:
                    workbook = load_workbook(report_path)
                    try:
                        self.assertEqual(workbook.sheetnames, [
                            "总览", "AutoML选型", "深度学习对比", "缺陷标签", "聚类画像", "特征重要性", "推理结果", "多分类评估",
                        ])
                        summary = dict(workbook["总览"].iter_rows(min_row=2, values_only=True))
                        self.assertEqual(summary["规则集版本"], "report_v2")
                        self.assertEqual(summary["评估配置"], "cross_validation: 3 folds")
                        self.assertEqual(workbook["AutoML选型"].max_row, 2)
                        automl_headers = [
                            cell.value for cell in workbook["AutoML选型"][1]
                        ]
                        self.assertIn("算法家族", automl_headers)
                        self.assertIn("最佳参数", automl_headers)
                        self.assertIn("完成试验", automl_headers)
                        self.assertEqual(workbook["缺陷标签"].max_row, 25)
                        self.assertEqual(workbook["特征重要性"].max_row, len(run.feature_schema) + 1)
                        self.assertEqual(workbook["推理结果"].max_row, 25)
                        evaluation_labels = {
                            str(row[0])
                            for row in workbook["多分类评估"].iter_rows(min_row=2, values_only=True)
                            if row[0] is not None
                        }
                        self.assertIn("strong_splatter", evaluation_labels)
                        self.assertIn("weak_splatter", evaluation_labels)
                        self.assertEqual(sum(len(sheet._images) for sheet in workbook.worksheets), 5)
                    finally:
                        workbook.close()
                rule_set = db.query(SpotWeldQualityRuleSet).filter(
                    SpotWeldQualityRuleSet.run_id == run.id,
                ).one()
                self.assertEqual(rule_set.version, "report_v2")
                self.assertEqual(rule_set.thresholds["energy_dev_sigma"], 2.5)
                self.assertNotIn("energy_dev_abs", rule_set.thresholds)
                self.assertEqual(rule_set.thresholds["spatter_cluster_id"], 1)
            finally:
                db.close()
                engine.dispose()

    def test_annotation_progress_does_not_reload_each_committed_sample(self):
        with tempfile.TemporaryDirectory(prefix="quality-progress-query-test-") as directory:
            engine = create_engine(
                "sqlite://",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            Session = sessionmaker(bind=engine)
            Base.metadata.create_all(engine)
            db = Session()
            try:
                owner = User(username=f"quality-progress-{uuid.uuid4().hex}", password_hash="hash")
                db.add(owner)
                db.flush()
                project = Project(name="Quality progress", owner_id=owner.id)
                db.add(project)
                db.commit()
                artifacts = ArtifactService(db, LocalStorage(Path(directory) / "artifacts"))
                dataset = create_demo_quality_dataset(
                    db,
                    project_id=project.id,
                    row_count=24,
                    artifact_service=artifacts,
                )
                db.commit()
                run = create_quality_run_record(
                    db,
                    project_id=project.id,
                    user_id=owner.id,
                    dataset_artifact_id=dataset.id,
                    algorithm_ids=["gbdt"],
                    search_method="random",
                    max_trials=5,
                    time_budget=60,
                    artifact_service=artifacts,
                )
                db.commit()
                sample_selects = []

                def capture_sample_selects(_conn, _cursor, statement, _parameters, _context, _executemany):
                    if statement.lstrip().upper().startswith("SELECT") and "spot_weld_quality_samples" in statement:
                        sample_selects.append(statement)

                event.listen(engine, "before_cursor_execute", capture_sample_selects)
                try:
                    outcome = execute_quality_run(db, run.id, artifact_service=artifacts)
                finally:
                    event.remove(engine, "before_cursor_execute", capture_sample_selects)

                self.assertEqual(outcome.status, "completed")
                self.assertLessEqual(len(sample_selects), 2, sample_selects)
            finally:
                db.close()
                engine.dispose()

    def test_manual_quality_run_leaves_automatic_labels_empty_for_operator_review(self):
        with tempfile.TemporaryDirectory(prefix="quality-manual-test-") as directory:
            engine = create_engine(
                "sqlite://",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            Session = sessionmaker(bind=engine)
            Base.metadata.create_all(engine)
            db = Session()
            try:
                owner = User(username=f"quality-manual-{uuid.uuid4().hex}", password_hash="hash")
                db.add(owner)
                db.flush()
                project = Project(name="Quality manual", owner_id=owner.id)
                db.add(project)
                db.commit()
                artifacts = ArtifactService(db, LocalStorage(Path(directory) / "artifacts"))
                dataset = create_demo_quality_dataset(
                    db,
                    project_id=project.id,
                    row_count=24,
                    artifact_service=artifacts,
                )
                db.commit()
                run = create_quality_run_record(
                    db,
                    project_id=project.id,
                    user_id=owner.id,
                    dataset_artifact_id=dataset.id,
                    algorithm_ids=["gbdt"],
                    search_method="random",
                    max_trials=5,
                    time_budget=60,
                    artifact_service=artifacts,
                )
                run.input_fingerprint = {**(run.input_fingerprint or {}), "label_mode": "manual"}
                db.commit()

                self.assertEqual(execute_quality_run(db, run.id, artifact_service=artifacts).status, "completed")
                samples = db.query(SpotWeldQualitySample).filter(
                    SpotWeldQualitySample.run_id == run.id,
                ).all()
                self.assertTrue(samples)
                self.assertTrue(all(sample.automatic_label is None for sample in samples))
                self.assertTrue(all(not sample.rule_hits for sample in samples))
            finally:
                db.close()
                engine.dispose()

    def test_quality_run_persists_and_executes_selected_algorithm_families(self):
        with tempfile.TemporaryDirectory(prefix="quality-candidate-test-") as directory:
            engine = create_engine(
                "sqlite://",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            Session = sessionmaker(bind=engine)
            Base.metadata.create_all(engine)
            db = Session()
            try:
                owner = User(username=f"quality-candidates-{uuid.uuid4().hex}", password_hash="hash")
                db.add(owner)
                db.flush()
                project = Project(name="Quality candidates", owner_id=owner.id)
                db.add(project)
                db.commit()
                artifacts = ArtifactService(db, LocalStorage(Path(directory) / "artifacts"))
                dataset = create_demo_quality_dataset(
                    db,
                    project_id=project.id,
                    row_count=24,
                    artifact_service=artifacts,
                )
                db.commit()

                run = create_quality_run_record(
                    db,
                    project_id=project.id,
                    user_id=owner.id,
                    dataset_artifact_id=dataset.id,
                    algorithm_ids=["gbdt", "random_forest"],
                    search_method="random",
                    max_trials=5,
                    time_budget=60,
                    artifact_service=artifacts,
                )
                db.commit()

                self.assertEqual(run.input_fingerprint["search_contract"], "optuna_v1")
                self.assertEqual(run.input_fingerprint["algorithm_ids"], ["gbdt", "random_forest"])
                self.assertEqual(execute_quality_run(db, run.id, artifact_service=artifacts).status, "completed")
                db.refresh(run)
                self.assertEqual(
                    [item["algorithm_id"] for item in run.automl_results],
                    ["gbdt", "random_forest"],
                )
                self.assertEqual(run.statistics["search"]["method"], "random")
                self.assertEqual(run.statistics["modeling_progress"]["total_count"], 10)
            finally:
                db.close()
                engine.dispose()

    def test_approved_snapshot_trains_a_quality_model_and_writes_report_workbook(self):
        with tempfile.TemporaryDirectory(prefix="quality-training-test-") as directory:
            engine = create_engine(
                "sqlite://",
                connect_args={"check_same_thread": False},
                poolclass=StaticPool,
            )
            Session = sessionmaker(bind=engine)
            Base.metadata.create_all(engine)
            db = Session()
            try:
                owner = User(username=f"quality-training-{uuid.uuid4().hex}", password_hash="hash")
                db.add(owner)
                db.flush()
                project = Project(name="Quality training", owner_id=owner.id)
                db.add(project)
                db.commit()
                artifacts = ArtifactService(db, LocalStorage(Path(directory) / "artifacts"))
                dataset = create_demo_quality_dataset(
                    db,
                    project_id=project.id,
                    row_count=24,
                    artifact_service=artifacts,
                )
                db.commit()
                run = create_quality_run_record(
                    db,
                    project_id=project.id,
                    user_id=owner.id,
                    dataset_artifact_id=dataset.id,
                    cross_validation_enabled=False,
                    cross_validation_folds=None,
                    algorithm_ids=["gbdt"],
                    search_method="random",
                    max_trials=5,
                    time_budget=60,
                    artifact_service=artifacts,
                )
                db.commit()
                self.assertEqual(execute_quality_run(db, run.id, artifact_service=artifacts).status, "completed")
                samples = db.query(SpotWeldQualitySample).filter(SpotWeldQualitySample.run_id == run.id).all()
                labels = []
                for sample in samples:
                    sample.current_label = sample.automatic_label
                    sample.review_status = "approved"
                    labels.append({"sample_id": str(sample.id), "label": sample.current_label, "revision_id": None, "source": "approved"})
                counts = Counter(item["label"] for item in labels)
                self.assertGreaterEqual(min(counts.values()), 5)
                snapshot = SpotWeldLabelSnapshot(
                    project_id=project.id,
                    run_id=run.id,
                    created_by_id=owner.id,
                    name="approved-for-training",
                    labels=labels,
                    label_counts=dict(counts),
                )
                db.add(snapshot)
                db.commit()

                outcome = train_label_snapshot(db, snapshot.id, artifact_service=artifacts)

                self.assertEqual(outcome.model_library.params["label_snapshot_id"], str(snapshot.id))
                self.assertEqual(outcome.model_library.params["feature_version"], "report_v1")
                self.assertEqual(outcome.model_library.params["quality_run_id"], str(run.id))
                self.assertEqual(outcome.model_library.params["label_source"], "approved")
                self.assertEqual(outcome.model_library.backbone, "gbdt")
                self.assertEqual(outcome.model_library.params["algorithm_id"], "gbdt")
                self.assertEqual(outcome.model_library.params["algorithm_ids"], ["gbdt"])
                self.assertEqual(outcome.model_library.params["search_method"], "random")
                self.assertTrue(outcome.model_library.params["best_params"])
                self.assertIn("report", outcome.output_artifacts)
                with artifacts.materialize(outcome.output_artifacts["report"], project.id, expected_type="quality_report") as report_path:
                    workbook = load_workbook(report_path, read_only=True)
                    try:
                        self.assertEqual(workbook.sheetnames, [
                            "总览", "AutoML选型", "深度学习对比", "缺陷标签",
                            "聚类画像", "特征重要性", "推理结果", "多分类评估",
                        ])
                        summary = dict(workbook["总览"].iter_rows(min_row=2, values_only=True))
                        self.assertEqual(summary["标签来源"], "approved")
                        self.assertEqual(summary["质量运行评估配置"], "deterministic_holdout")
                        self.assertEqual(summary["快照训练评估配置"], "cross_validation: 5 folds")
                        self.assertEqual(summary["训练标签样本"], len(snapshot.labels))
                        self.assertNotIn("已审核样本", summary)
                        self.assertEqual(workbook["AutoML选型"].max_row, 2)
                        snapshot_names = [
                            str(row[0])
                            for row in workbook["AutoML选型"].iter_rows(min_row=2, values_only=True)
                        ]
                        self.assertFalse(any(name.startswith("AutoML(") or name.startswith("MLP_") for name in snapshot_names))
                    finally:
                        workbook.close()
            finally:
                db.close()
                engine.dispose()


if __name__ == "__main__":
    unittest.main()
